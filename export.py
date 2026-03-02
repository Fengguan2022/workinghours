from collections import defaultdict
from datetime import datetime
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

import database


def export_to_excel(filepath: str, start_date: str, end_date: str, employee_id: Optional[int] = None):
    entries = database.get_entries_in_range(start_date, end_date, employee_id)
    overtime_threshold = database.get_overtime_threshold()

    # Build employee name cache
    emp_cache = {}
    for entry in entries:
        if entry.employee_id not in emp_cache:
            emp = database.get_employee_by_id(entry.employee_id)
            emp_cache[entry.employee_id] = emp.name if emp else "Unknown"

    wb = Workbook()

    _build_daily_summary(wb, entries, emp_cache, overtime_threshold)
    _build_raw_records(wb, entries, emp_cache)
    _build_monthly_summary(wb, entries, emp_cache, overtime_threshold)

    wb.save(filepath)


def _header_style():
    return Font(bold=True, color="FFFFFF"), PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")


def _build_daily_summary(wb: Workbook, entries, emp_cache, overtime_threshold):
    ws = wb.active
    ws.title = "Daily Summary"

    headers = ["Employee", "Date", "Clock In", "Clock Out", "Total Hours", "Overtime"]
    font, fill = _header_style()
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = font
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")

    # Group entries by employee and date
    by_emp_date = defaultdict(list)
    for e in entries:
        dt = datetime.strptime(e.timestamp, "%Y-%m-%d %H:%M:%S")
        key = (e.employee_id, dt.strftime("%Y-%m-%d"))
        by_emp_date[key].append(e)

    row = 2
    for (emp_id, date_str), day_entries in sorted(by_emp_date.items(), key=lambda x: (emp_cache.get(x[0][0], ""), x[0][1])):
        day_entries.sort(key=lambda e: e.timestamp)

        # Pair IN/OUT
        i = 0
        total_hours = 0.0
        pairs = []
        while i < len(day_entries):
            if day_entries[i].event_type == "IN":
                clock_in = datetime.strptime(day_entries[i].timestamp, "%Y-%m-%d %H:%M:%S")
                clock_out = None
                if i + 1 < len(day_entries) and day_entries[i + 1].event_type == "OUT":
                    clock_out = datetime.strptime(day_entries[i + 1].timestamp, "%Y-%m-%d %H:%M:%S")
                    hours = (clock_out - clock_in).total_seconds() / 3600
                    total_hours += hours
                    pairs.append((clock_in, clock_out))
                    i += 2
                else:
                    pairs.append((clock_in, None))
                    i += 1
            else:
                i += 1

        is_overtime = total_hours > overtime_threshold

        if pairs:
            for ci, co in pairs:
                ws.cell(row=row, column=1, value=emp_cache.get(emp_id, "Unknown"))
                ws.cell(row=row, column=2, value=date_str)
                ws.cell(row=row, column=3, value=ci.strftime("%H:%M:%S"))
                ws.cell(row=row, column=4, value=co.strftime("%H:%M:%S") if co else "Still IN")
                ws.cell(row=row, column=5, value=round(total_hours, 2))
                ws.cell(row=row, column=6, value="Y" if is_overtime else "N")
                row += 1
        else:
            ws.cell(row=row, column=1, value=emp_cache.get(emp_id, "Unknown"))
            ws.cell(row=row, column=2, value=date_str)
            ws.cell(row=row, column=5, value=0)
            ws.cell(row=row, column=6, value="N")
            row += 1

    # Auto-width
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = max(max_len + 2, 12)


def _build_raw_records(wb: Workbook, entries, emp_cache):
    ws = wb.create_sheet("Raw Records")

    headers = ["Employee", "Timestamp", "Event Type"]
    font, fill = _header_style()
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = font
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")

    for i, entry in enumerate(sorted(entries, key=lambda e: e.timestamp), start=2):
        ws.cell(row=i, column=1, value=emp_cache.get(entry.employee_id, "Unknown"))
        ws.cell(row=i, column=2, value=entry.timestamp)
        ws.cell(row=i, column=3, value=entry.event_type)

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = max(max_len + 2, 12)


def _build_monthly_summary(wb: Workbook, entries, emp_cache, overtime_threshold):
    ws = wb.create_sheet("Monthly Summary")

    headers = ["Employee", "Month", "Total Hours", "Overtime Hours", "Days Worked"]
    font, fill = _header_style()
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = font
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")

    # Group by employee and month
    by_emp_month = defaultdict(lambda: defaultdict(list))
    for e in entries:
        dt = datetime.strptime(e.timestamp, "%Y-%m-%d %H:%M:%S")
        month_key = dt.strftime("%Y-%m")
        by_emp_month[e.employee_id][month_key].append(e)

    row = 2
    for emp_id in sorted(by_emp_month, key=lambda eid: emp_cache.get(eid, "")):
        for month_key in sorted(by_emp_month[emp_id]):
            month_entries = sorted(by_emp_month[emp_id][month_key], key=lambda e: e.timestamp)

            # Group by day to calculate daily hours and overtime
            by_day = defaultdict(list)
            for e in month_entries:
                dt = datetime.strptime(e.timestamp, "%Y-%m-%d %H:%M:%S")
                by_day[dt.strftime("%Y-%m-%d")].append(e)

            total_hours = 0.0
            overtime_hours = 0.0
            days_worked = set()

            for day_str, day_entries in by_day.items():
                day_entries.sort(key=lambda e: e.timestamp)
                day_hours = 0.0
                i = 0
                while i < len(day_entries):
                    if day_entries[i].event_type == "IN":
                        ci = datetime.strptime(day_entries[i].timestamp, "%Y-%m-%d %H:%M:%S")
                        if i + 1 < len(day_entries) and day_entries[i + 1].event_type == "OUT":
                            co = datetime.strptime(day_entries[i + 1].timestamp, "%Y-%m-%d %H:%M:%S")
                            day_hours += (co - ci).total_seconds() / 3600
                            i += 2
                        else:
                            i += 1
                    else:
                        i += 1

                if day_hours > 0:
                    days_worked.add(day_str)
                total_hours += day_hours
                if day_hours > overtime_threshold:
                    overtime_hours += day_hours - overtime_threshold

            ws.cell(row=row, column=1, value=emp_cache.get(emp_id, "Unknown"))
            ws.cell(row=row, column=2, value=month_key)
            ws.cell(row=row, column=3, value=round(total_hours, 2))
            ws.cell(row=row, column=4, value=round(overtime_hours, 2))
            ws.cell(row=row, column=5, value=len(days_worked))
            row += 1

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = max(max_len + 2, 12)
